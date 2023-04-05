#----------PYTHON REUSABLE - UPDATING EXCEL---------------
#-----------------------------------------
def string_exist_check(source_str, checking_str):
    result = checking_str in source_str
    #print(x)
    return result

def create_columns_if_not_exists(ws, columns):
    """
    Creates the columns if they do not exist.

    parameters:
        `ws` (object): Worksheet object.
        `columns` (list): List of column headers.

    Returns:
        ws (Worksheet): Worksheet object.
        colss (dict): Dictionary of column names and their respective column numbers.
    """
    colss = {}
    for column in columns:
        for col in ws.iter_cols():
            if column.lower() == col[0].value.lower():
                colss[column] = col[0].column-1
                break
        else:
            ws.insert_cols(ws.max_column+1)
            statuscell = ws.cell(row=1, column=ws.max_column+1)
            statuscell.value = column
            colss[column] = statuscell.column-1
    return ws, colss

def update_excel_value(filename, criteriacolumnheader, criteriacellvalue, statuscolumnheader, statusvalue):
    """
    Updates the excel file with the given values.

    Returns:
        True if the value is updated.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    ws = wb['Invoice_Sheet']
    ws, statuscolumn = create_columns_if_not_exists(ws, [statuscolumnheader])
    for col in ws.iter_cols():
        if criteriacolumnheader == col[0].value:
            for row in ws.iter_rows():
                if str(criteriacellvalue) == str(row[col[0].column-1].value):
                    row[statuscolumn[statuscolumnheader]].value = statusvalue
                    wb.save(filename)
                    return True
# update_excel_value(r"testdata\Automation Test Case.xlsx","TEST CASE ID","TC_01_Portal Launching","Status","pass")